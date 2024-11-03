import asyncio
import os
import tempfile

import pandas as pd
import pytest
from openpyxl import load_workbook
from rich import print

from edgar import Filing, Company
from edgar.financials import (Financials, MultiFinancials, BalanceSheet, CashFlowStatement,
                              IncomeStatement, StatementOfChangesInEquity, StatementOfComprehensiveIncome)
from edgar.xbrl import XBRLData, XBRLInstance, Statement, Statements, get_xbrl_object
from edgar.xbrl.xbrldata import get_primary_units, get_unit_divisor


@pytest.fixture(scope='module')
def apple_xbrl():
    filing: Filing = Filing(company='Apple Inc.', cik=320193, form='10-K', filing_date='2023-11-03',
                            accession_no='0000320193-23-000106')
    return asyncio.run(XBRLData.from_filing(filing))


@pytest.fixture()
def teradyne_xbrl():
    f = Filing(company='TERADYNE, INC', cik=97210, form='10-Q', filing_date='2024-08-02',
               accession_no='0000950170-24-089858')
    return f.xbrl()


@pytest.fixture(scope='module')
def tesla_xbrl():
    filing: Filing = Filing(company='Tesla, Inc.', cik=1318605, form='10-Q', filing_date='2024-07-24',
                            accession_no='0001628280-24-032662')
    return get_xbrl_object(filing)


@pytest.fixture(scope='module')
def netflix_xbrl():
    filing: Filing = Filing(company='NETFLIX INC', cik=1065280, form='10-Q', filing_date='2024-04-22',
                            accession_no='0001065280-24-000128')
    return asyncio.run(XBRLData.from_filing(filing))


@pytest.fixture(scope='module')
def crowdstrike_xbrl():
    filing = Filing(company='CrowdStrike Holdings, Inc.', cik=1535527, form='10-K', filing_date='2024-03-07',
                    accession_no='0001535527-24-000007')
    return asyncio.run(XBRLData.from_filing(filing))


@pytest.fixture(scope='module')
def orcl_xbrl():
    filing = Filing(company='ORACLE CORP', cik=1341439, form='10-K', filing_date='2024-06-20',
                    accession_no='0000950170-24-075605')
    return asyncio.run(XBRLData.from_filing(filing))


@pytest.fixture(scope='module')
def msft_xbrl():
    filing = Filing(company='MICROSOFT CORP', cik=789019, form='10-K', filing_date='2023-07-27',
                    accession_no='0000950170-23-035122')
    return XBRLData.extract(filing)


@pytest.fixture(scope='module')
def gd_xbrl():
    filing = Filing(company='GENERAL DYNAMICS CORP', cik=40533, form='10-Q', filing_date='2024-07-24',
                    accession_no='0000040533-24-000035')
    return XBRLData.extract(filing)


@pytest.fixture(scope='module')
def pfizer_xbrl():
    filing = Filing(company='PFIZER INC', cik=78003, form='10-K', filing_date='2024-02-22',
                    accession_no='0000078003-24-000039')
    return get_xbrl_object(filing)


@pytest.mark.asyncio
async def test_get_shareholder_equity_statement_for_10K(apple_xbrl):
    instance = apple_xbrl.instance
    pr = apple_xbrl.presentation
    sd = apple_xbrl.get_statement_definition('CONSOLIDATEDSTATEMENTSOFSHAREHOLDERSEQUITY')
    lic = sd._find_line_items_container(pr.roles['http://www.apple.com/role/CONSOLIDATEDSTATEMENTSOFSHAREHOLDERSEQUITY'])
    statement: Statement = Financials(apple_xbrl).get_statement_of_changes_in_equity()
    statement.print_structure()
    assert statement
    print(statement)
    print(statement.data)
    #shareholders_equity = statement.get_concept(concept='us-gaap_StockholdersEquity')
    #assert shareholders_equity.value == {'2023': '62146000000'}


def test_get_statement_definition_line_item_root(apple_xbrl):
    sd = apple_xbrl.get_statement_definition('CONSOLIDATEDSTATEMENTSOFCASHFLOWS')

    statement = apple_xbrl.get_statement('CONSOLIDATEDSTATEMENTSOFCASHFLOWS')
    o = statement.__rich__()
    print(o)


@pytest.mark.asyncio
async def test_statement_get_concept_value(apple_xbrl):
    statement: Statement = Financials(apple_xbrl).get_statement_of_changes_in_equity()
    concept = statement.data.query("concept == 'us-gaap_NetIncomeLoss' ")
    print()
    assert concept['2023'][0] == '96995000000'
    assert concept['2022'][0] == '99803000000'
    assert concept['2021'][0] == '94680000000'


def test_get_balance_sheet(apple_xbrl):
    balance_sheet: Statement = Financials(apple_xbrl).get_balance_sheet()
    print(balance_sheet)
    assert balance_sheet.periods == ['2023', '2022']


def test_cover_page_aapl(apple_xbrl):
    cover_page = apple_xbrl.get_statement('CoverPage')
    print(cover_page)
    assert cover_page is not None
    registrant_name =  cover_page.get_concept(concept='dei_EntityRegistrantName')
    assert registrant_name.value['2023'] == 'Apple Inc.'


def test_get_concept_from_statement(tesla_xbrl):
    financials: Financials = Financials(tesla_xbrl)
    income_statement = financials.get_income_statement()
    assert income_statement.durations == {'3 months', '6 months'}
    concept = income_statement.data.query("concept=='us-gaap_RevenueFromContractWithCustomerExcludingAssessedTax'")
    print()
    print(concept)
    #assert concept
    #assert concept.value == {'Jun 30, 2023': '24927000000', 'Jun 30, 2024': '25500000000'}


def test_get_concept_using_label(apple_xbrl):
    cover_page: Statement = apple_xbrl.get_statement('CoverPage', include_concept=True)
    assert cover_page is not None
    fact = cover_page.get_concept(label='Entity Registrant Name')
    assert fact.value['2023'] == 'Apple Inc.'
    assert fact.name == 'dei_EntityRegistrantName'


def test_statements_property(apple_xbrl):
    statements: Statements = apple_xbrl.statements
    assert len(statements) == 78
    assert 'CoverPage' in statements


def test_10Q_filings_have_quarterly_dates(netflix_xbrl):
    balance_sheet: Statement = Financials(netflix_xbrl).get_balance_sheet()
    print(balance_sheet)
    assert balance_sheet.periods == ['Mar 31, 2024', 'Dec 31, 2023']
    for name in netflix_xbrl.list_statement_definitions():
        print(name)


@pytest.mark.asyncio
async def test_labels_for_orcl_10K(orcl_xbrl):
    financials: Financials = Financials(orcl_xbrl)
    balance_sheet = financials.get_balance_sheet()
    print(balance_sheet)
    #assert not balance_sheet.labels[0].startswith('us-gaap_')


@pytest.mark.asyncio
async def test_labels_for_msft_10K(msft_xbrl):
    financials: Financials = Financials(msft_xbrl)
    balance_sheet = financials.get_balance_sheet()
    first_label = balance_sheet.data.index[0]
    print(balance_sheet)
    assert first_label == 'Cash and cash equivalents'
    assert not '_' in balance_sheet.labels[0]


def test_get_all_dimensions(apple_xbrl):
    instance: XBRLInstance = apple_xbrl.instance
    dimensions = instance.get_all_dimensions()
    assert {
               'us-gaap:AwardTypeAxis',
               'us-gaap:ConcentrationRiskByTypeAxis',
               'us-gaap:LongtermDebtTypeAxis',
               'us-gaap:FairValueByFairValueHierarchyLevelAxis',
               'us-gaap:AntidilutiveSecuritiesExcludedFromComputationOfEarningsPerShareByAntidilutiveSecuritiesAxis',
           } & dimensions


def test_get_dimension_values(apple_xbrl):
    instance: XBRLInstance = apple_xbrl.instance
    values = instance.get_dimension_values('us-gaap:LongtermDebtTypeAxis')
    assert values == ['aapl:FixedRateNotesMember']
    assert instance.get_dimension_values('us-gaap:NonExisting') == []


def test_query_facts_by_dimension_value(apple_xbrl):
    instance: XBRLInstance = apple_xbrl.instance
    facts = instance.query_facts(dimensions={'ecd:IndividualAxis': 'aapl:DeirdreOBrienMember'})
    assert facts['ecd:IndividualAxis'].drop_duplicates().tolist() == ['aapl:DeirdreOBrienMember']


def test_query_facts_by_dimension_values_as_list(apple_xbrl):
    instance: XBRLInstance = apple_xbrl.instance
    facts = instance.query_facts(
        dimensions={'ecd:IndividualAxis': ['aapl:DeirdreOBrienMember', 'aapl:JeffWilliamsMember']})
    assert len(facts) == 12
    assert facts['ecd:IndividualAxis'].drop_duplicates().tolist() == ['aapl:DeirdreOBrienMember',
                                                                      'aapl:JeffWilliamsMember']


def test_statements_contain_dimension_axis_values(apple_xbrl):
    instance: XBRLInstance = apple_xbrl.instance
    dimensions = instance.dimensions
    axis = dimensions['srt:ProductOrServiceAxis']
    assert axis.list_members() == ['us-gaap:ProductMember',
                                   'us-gaap:ServiceMember',
                                   'aapl:IPhoneMember',
                                   'aapl:MacMember',
                                   'aapl:IPadMember',
                                   'aapl:WearablesHomeandAccessoriesMember']


def test_query_facts_with_empty_dimensions(apple_xbrl):
    instance: XBRLInstance = apple_xbrl.instance
    facts = instance.query_facts(dimensions={})

    assert all([col not in instance.dimension_columns for col in facts.columns])


def test_xbrl_instance_dimensions(apple_xbrl):
    instance: XBRLInstance = apple_xbrl.instance
    print(instance.dimensions)


def test_xbrl_presentation_role_with_almost_duplicate_name(apple_xbrl):
    statement = apple_xbrl.statements.get('RevenueDeferredRevenueExpectedTimingofRealizationDetails_1')
    assert statement is None


def test_list_standard_statements(apple_xbrl):
    financials: Financials = Financials(apple_xbrl)
    statements = financials.list_standard_statements()
    assert statements == ['Cover Page', 'Consolidated Balance Sheets', 'Income Statements',
                          'Consolidated Statement of Cash Flows', 'Consolidated Statement of Shareholders Equity',
                          'Comprehensive Income Statement']
    print()
    print(financials)


@pytest.mark.asyncio
async def test_xbrl_financials_using_non_standard_filing_like_crowdstrike(crowdstrike_xbrl):
    financials: Financials = Financials(crowdstrike_xbrl)

    balance_sheet = financials.get_balance_sheet()
    assert balance_sheet

    comprehensive_income_statement = financials.get_statement_of_comprehensive_income()
    assert comprehensive_income_statement

    cashflow_statement = financials.get_cash_flow_statement()
    assert cashflow_statement

    equity_statement = financials.get_statement_of_changes_in_equity()
    assert equity_statement

    cover_page = financials.get_cover_page()
    assert cover_page

    income_statement = financials.get_income_statement()
    assert income_statement


def test_extract_financials_from_filing(gd_xbrl):
    financials: Financials = Financials(gd_xbrl)
    assert financials


def test_quarterly_data_extracted_correctly(gd_xbrl):
    assert gd_xbrl
    financials: Financials = Financials(gd_xbrl)
    income_statement: Statement = financials.get_income_statement()

    # This is wrong because the values are dimensioned
    # There is a concept for us-gaap:CostOfGoodsAndServicesSold
    # a dimension for each of the product and services and {'srt:ProductOrServiceAxis': 'us-gaap:ProductMember'}
    concept = income_statement.data.query("concept=='us-gaap_CostOfGoodsAndServicesSold'")
    assert len(concept)==3


def test_3month_extracted_correctly_for_appfolio():
    filing = Filing(company='APPFOLIO INC', cik=1433195,
                    form='10-Q', filing_date='2023-10-27',
                    accession_no='0001433195-23-000126')
    xbrl_data = get_xbrl_object(filing)
    # The default income statement is for three months
    income_statement = xbrl_data.get_statement("CONDENSEDCONSOLIDATEDSTATEMENTSOFOPERATIONS")

    data = income_statement.data
    assert data.loc['Revenue', 'Sep 30, 2023'] == '165440000'
    assert data.loc['Revenue', 'Sep 30, 2022'] == '125079000'

    assert data.loc['Net income (loss)', 'Sep 30, 2023'] == '26445000'
    assert data.loc['Net income (loss)', 'Sep 30, 2022'] == '-4162000'


def test_six_month_data_extracted_correctly():
    filing = Filing(company='APPFOLIO INC', cik=1433195, form='10-Q', filing_date='2024-07-26',
                    accession_no='0001433195-24-000106')
    xbrl_data = get_xbrl_object(filing)
    # The nine month income statement
    income_statement = xbrl_data.get_statement("CONDENSEDCONSOLIDATEDSTATEMENTSOFOPERATIONS",
                                               duration="6 months")
    assert income_statement
    print(income_statement)
    data = income_statement.data
    assert data.loc['Revenue', 'Jun 30, 2024'] == '384805000'
    assert data.loc['Revenue', 'Jun 30, 2023'] == '283175000'


def test_nine_month_data_extracted_correctly():
    filing = Filing(company='APPFOLIO INC', cik=1433195,
                    form='10-Q', filing_date='2023-10-27',
                    accession_no='0001433195-23-000126')
    xbrl_data = get_xbrl_object(filing)
    # The nine month income statement
    income_statement = xbrl_data.get_statement("CONDENSEDCONSOLIDATEDSTATEMENTSOFOPERATIONS",
                                               duration="9 months")
    assert income_statement
    print(income_statement)
    data = income_statement.data
    assert data.loc['Revenue', 'Sep 30, 2023'] == '448615000'
    assert data.loc['Revenue', 'Sep 30, 2022'] == '347825000'


def test_get_concepts_for_label(gd_xbrl):
    concept = 'us-gaap_OperatingIncomeLoss'
    assert gd_xbrl.get_concept_for_label("Operating Income (Loss)") == concept
    assert gd_xbrl.get_concept_for_label("Operating Earnings") == concept
    assert gd_xbrl.get_concept_for_label("Operating earnings") == concept

    assert gd_xbrl.get_labels_for_concept(concept) == {"label": "Operating Income (Loss)",
                                                       "terseLabel": "Operating Earnings",
                                                       'totalLabel': "Operating earnings"}

    roles = ['http://www.generaldynamics.com/role/ConsolidatedStatementofEarningsUnaudited',
             'http://www.generaldynamics.com/role/RevenueImpactofAdjustmentsinContractEstimatesDetails',
             'http://www.generaldynamics.com/role/SegmentInformationSummaryofFinancialInformationDetails']
    # Roles for label
    assert gd_xbrl.get_roles_for_label("Operating Income (Loss)") == roles
    statements = gd_xbrl.list_statements_for_label("Operating Income (Loss)")
    print(statements)

    statement = gd_xbrl.get_statement(statements[0])
    assert statement


def test_handle_financials_with_only_document_and_entity_definition():
    # This filing only has 'DocumentAndEntityInformation'
    filing = Filing(form='10-K/A', filing_date='2024-07-26', company='Swiftmerge Acquisition Corp.',
                    cik=1845123, accession_no='0001013762-24-001580')
    xbrl_data: XBRLData = XBRLData.extract(filing)
    statement_definitions = xbrl_data.list_statement_definitions()
    assert statement_definitions == ['DocumentAndEntityInformation']
    financials = Financials(xbrl_data)
    assert financials
    assert financials.get_cover_page()
    assert not financials.get_balance_sheet()
    assert not financials.get_income_statement()
    assert not financials.get_cash_flow_statement()
    assert not financials.get_statement_of_changes_in_equity()
    assert not financials.get_statement_of_comprehensive_income()
    assert financials.get_cover_page()


def test_get_dataframe_from_statement(apple_xbrl):
    financials: Financials = Financials(apple_xbrl)
    balance_sheet = financials.get_balance_sheet()
    assert balance_sheet.get_dataframe(include_concept=False).columns.tolist() == ['2023', '2022']
    assert balance_sheet.get_dataframe(include_concept=True).columns.tolist() == ['2023', '2022', 'concept']
    assert balance_sheet.get_dataframe(include_format=True).columns.tolist() == ['2023', '2022', 'concept', 'level', 'decimals', 'style']
    assert balance_sheet.get_dataframe(include_concept=True, include_format=True).columns.tolist() == ['2023', '2022',
                                                                                                       'concept',
                                                                                                       'level',
                                                                                                       'decimals',
                                                                                                       'style',
                                                                                                       ]


def test_get_primary_units():
    # Test case for Millions
    assert get_primary_units(1_000_000) == "Millions"

    # Test case for Hundreds of Thousands
    assert get_primary_units(100_000) == "Hundreds of Thousands"

    # Test case for Thousands
    assert get_primary_units(1_000) == "Thousands"

    # Test case for Hundreds
    assert get_primary_units(100) == "Hundreds"

    # Test case for Tens
    assert get_primary_units(10) == "Tens"

    # Test case for Units (default)
    assert get_primary_units(1) == "Units"


def test_get_unit_divisor():
    # Test case with -6 and -3 in the column
    data1 = {'decimals': ['-6', '-6', '-3', '-3', 'INF', '0']}
    df1 = pd.DataFrame(data1)
    assert get_unit_divisor(df1) == 1000  # 10 ** 3

    # Test case with only -3 in the column
    data2 = {'decimals': ['-3', '-3', 'INF', '0']}
    df2 = pd.DataFrame(data2)
    assert get_unit_divisor(df2) == 1000  # 10 ** 3

    # Test case with only -6 in the column
    data3 = {'decimals': ['-6', 'INF', '0']}
    df3 = pd.DataFrame(data3)
    assert get_unit_divisor(df3) == 1000000  # 10 ** 6

    # Test case with no negative decimals
    data4 = {'decimals': ['INF', '0', '2']}
    df4 = pd.DataFrame(data4)
    assert get_unit_divisor(df4) == 1  # Default to no scaling

    # Test case with mixed positive and negative decimals
    data5 = {'decimals': ['-2', '-4', 'INF', '2', '0']}
    df5 = pd.DataFrame(data5)
    assert get_unit_divisor(df5) == 100  # 10 ** 2


def test_financials_extract_from_filing():
    filing = Filing(form='10-K', filing_date='2024-07-26', company='Swiftmerge Acquisition Corp.',
                    cik=1845123, accession_no='0001013762-24-001580')
    financials = Financials.extract(filing)
    assert financials
    assert isinstance(financials, Financials)


def test_financials_extract_filing():
    filing = Filing(form='10-K', filing_date='2024-07-26', company='Swiftmerge Acquisition Corp.',
                    cik=1845123, accession_no='0001013762-24-001580')
    financials = Financials.extract(filing)
    assert financials
    assert isinstance(financials, Financials)


def test_get_nonfinancial_statement(apple_xbrl):
    statements = apple_xbrl.statements
    assert len(statements) == 78

    # get the cover page
    cover_page: Statement = statements.get('CoverPage')
    assert cover_page
    assert isinstance(cover_page, Statement)
    assert cover_page.get_concept('EntityRegistrantName').value['2023'] == 'Apple Inc.'
    # Make sure that repr works
    _repr_ = repr(cover_page)
    print(_repr_)
    assert 'Document Type' in _repr_

    for statement in statements:
        print(statement.name)
        assert repr(statement)


def test_get_statement_from_statement_by_int_index(apple_xbrl):
    statements = apple_xbrl.statements
    assert len(statements)

    statement: Statement = statements[0]
    assert statement
    assert statement.display_name == "Cover"


def test_pfizer_current_assets(pfizer_xbrl):
    financials = Financials(pfizer_xbrl)
    balance_sheet = financials.get_balance_sheet()

    data = balance_sheet.data

    assert data.loc['Goodwill', '2023'] == '67783000000'
    assert data.loc['Goodwill', '2022'] == '51375000000'
    assert data.loc['Total current assets', '2023'] == '43333000000'
    assert data.loc['Total current assets', '2022'] == '51259000000'

    assert data.loc['Total assets', '2023'] == '227000000000'
    assert data.loc['Total assets', '2022'] == '197000000000'


def test_get_correct_value_of_marketable_securities(apple_xbrl):
    financials = Financials(apple_xbrl)
    balance_sheet = financials.get_balance_sheet()
    assert balance_sheet.get_concept('us-gaap_MarketableSecuritiesCurrent').value.get('2022') == '24658000000'
    assert balance_sheet.get_concept('us-gaap_MarketableSecuritiesCurrent').value.get('2023') == '31590000000'


@pytest.fixture
def temp_excel_file():
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
        yield tmp.name
    os.unlink(tmp.name)


def test_statement_to_excel_file(apple_xbrl, temp_excel_file):
    financials = Financials(apple_xbrl)
    balance_sheet = financials.get_balance_sheet()

    balance_sheet.to_excel(temp_excel_file)

    # Verify the file was created
    assert os.path.exists(temp_excel_file)

    # Load the Excel file and check its contents
    wb = load_workbook(temp_excel_file)
    assert len(wb.sheetnames) == 1
    sheet = wb.active

    # Check if the first row (header) matches the DataFrame columns
    df_columns = balance_sheet.get_dataframe(include_concept=True).columns.tolist()
    if 'label' != df_columns[0]:
        df_columns = ['label'] + df_columns
    excel_header = [cell.value for cell in sheet[1]]
    assert excel_header == df_columns


def test_statement_to_excel_writer(apple_xbrl, temp_excel_file):
    financials = Financials(apple_xbrl)
    balance_sheet = financials.get_balance_sheet()
    income_statement = financials.get_income_statement()

    with pd.ExcelWriter(temp_excel_file, engine='xlsxwriter') as writer:
        balance_sheet.to_excel(excel_writer=writer, include_format=True, include_concept=True)
        income_statement.to_excel(excel_writer=writer, include_format=False, include_concept=False)

    # Verify the file was created
    assert os.path.exists(temp_excel_file)

    # Load the Excel file and check its contents
    wb = load_workbook(temp_excel_file)
    assert len(wb.sheetnames) == 2

    # Check balance sheet
    bs_sheet = wb['CONSOLIDATEDBALANCESHEETS']
    bs_df = balance_sheet.get_dataframe(include_format=True, include_concept=True)
    bs_excel_header = [cell.value for cell in bs_sheet[1]]
    assert bs_excel_header == ['label'] + bs_df.columns.tolist()

    # Check income statement
    is_sheet = wb['CONSOLIDATEDSTATEMENTSOFOPERATIONS'[:31]]
    is_df = income_statement.get_dataframe(include_format=False, include_concept=False)
    is_excel_header = [cell.value for cell in is_sheet[1]]
    assert is_excel_header == ['label'] + is_df.columns.tolist()


def test_multi_financials_values():
    company = Company("AAPL", include_old_filings=False)
    filings = company.get_filings(form="10-K").filter(filing_date="2020-01-01:2024-03-01").latest(3)
    print(filings)
    multi_financials = MultiFinancials(filings)

    balance_sheet: Statement = multi_financials.get_balance_sheet()
    assert balance_sheet.get_concept('us-gaap_CashAndCashEquivalentsAtCarryingValue').value == {'2023': '29965000000',
                                                                                                '2022': '23646000000',
                                                                                                '2021': '34940000000',
                                                                                                '2020': '38016000000'}
    std_balance_sheet = multi_financials.get_balance_sheet(standard=True)
    assert std_balance_sheet is not None
    assert std_balance_sheet.get_concept('us-gaap_CashAndCashEquivalentsAtCarryingValue').value == {
        '2023': '29965000000',
        '2022': '23646000000',
        '2021': '34940000000',
        '2020': '38016000000'}

    income_statement: Statement = multi_financials.get_income_statement()
    assert income_statement.get_concept('us-gaap_NetIncomeLoss').value == {'2023': '96995000000',
                                                                           '2022': '99803000000',
                                                                           '2021': '94680000000',
                                                                           '2020': '57411000000',
                                                                           '2019': '55256000000'}
    # Cash flow statement
    cash_flow: Statement = multi_financials.get_cash_flow_statement()
    cashflow_values = cash_flow.get_concept('us-gaap_NetCashProvidedByUsedInOperatingActivities').value
    assert cashflow_values == {'2023': '110543000000',
                               '2022': '122151000000',
                               '2021': '104038000000',
                               '2020': '80674000000',
                               '2019': '69391000000'}

    company = Company("TSLA", include_old_filings=False)
    filings = company.get_filings(form="10-K").latest(3)
    multi_financials = MultiFinancials(filings)
    balance_sheet = multi_financials.get_balance_sheet()
    assert balance_sheet.data.columns.tolist() == ['2023', '2022', '2021', '2020', 'concept',
                                                   'level',
                                                   'decimals',
                                                   'style'
                                                   ]
    # Check that the concepts are unique
    assert balance_sheet.data.concept.nunique() == len(balance_sheet.data)
    # Test concept values
    assert balance_sheet.get_concept('us-gaap_CashAndCashEquivalentsAtCarryingValue').value == {
        '2023': '16398000000', '2022': '16253000000', '2021': '17576000000', '2020': '19384000000'}

    income_statement = multi_financials.get_income_statement()
    assert income_statement.data.columns.tolist() == ['2023', '2022', '2021', '2020', '2019', 'concept',
                                                      'segment', 'level', 'decimals', 'style']
    # Get the concept for a multifinancial
    netincome = income_statement.get_concept('us-gaap_NetIncomeLoss')
    assert income_statement.data.concept.nunique() == len(income_statement.data)
    assert netincome.value == {'2023': '14997000000',
                               '2022': '12556000000',
                               '2021': '5519000000',
                               '2020': '721000000',
                               '2019': '-862000000'}

    # Standardized balance sheet statement
    balance_sheet = multi_financials.get_balance_sheet(standard=True)
    assert set(balance_sheet.concepts).issubset({concept.concept for concept in BalanceSheet.concepts})


@pytest.mark.asyncio
async def test_multifinanancials_async():
    company = Company("AAPL", include_old_filings=False)
    filings = company.get_filings(form="10-K").latest(3)
    multi_financials = await MultiFinancials.extract_async(filings)
    assert multi_financials
    assert isinstance(multi_financials, MultiFinancials)
    balance_sheet = multi_financials.get_balance_sheet()
    columns = balance_sheet.get_dataframe().columns.tolist()
    assert len(columns) > 2


def test_apple_cashflow_correct_negative_values(apple_xbrl):
    financials = Financials(apple_xbrl)
    cash_flow = financials.get_cash_flow_statement()
    print(cash_flow)
    cashflow_values = cash_flow.get_concept('us-gaap_PaymentsForRepurchaseOfCommonStock').value
    assert cashflow_values == {'2023': '-77550000000',
                               '2022': '-89402000000',
                               '2021': '-85971000000'}

    assert cash_flow.get_concept('us-gaap_IncreaseDecreaseInInventories').value == {
        '2023': '-1618000000',
        '2022': '-1484000000',
        '2021': '-2642000000'
    }
    # All positive values
    assert cash_flow.get_concept('us-gaap_ShareBasedCompensation').value == {
        '2023': '10833000000',
        '2022': '9038000000',
        '2021': '7906000000'}


def test_standardized_statements(apple_xbrl):
    aapl_financials = Financials(apple_xbrl)

    # Test balance sheet
    balance_sheet = aapl_financials.get_balance_sheet(standard=True)
    assert balance_sheet is not None
    std_concepts = {concept.concept for concept in BalanceSheet.concepts}
    actual_concepts = set(balance_sheet.data.reset_index()['concept'])
    assert std_concepts.issuperset(
        actual_concepts), f"Missing concepts in balance sheet: {std_concepts - actual_concepts}"

    # Test income statement
    income_statement = aapl_financials.get_income_statement(standard=True)
    assert income_statement is not None
    std_concepts = {concept.concept for concept in IncomeStatement.concepts}
    actual_concepts = set(income_statement.data.reset_index()['concept'])
    assert std_concepts.issuperset(
        actual_concepts), f"Missing concepts in income statement: {std_concepts - actual_concepts}"

    # Test cash flow statement
    cash_flow = aapl_financials.get_cash_flow_statement(standard=True)
    assert cash_flow is not None
    std_concepts = {concept.concept for concept in CashFlowStatement.concepts}
    actual_concepts = set(cash_flow.data.reset_index()['concept'])
    assert std_concepts.issuperset(
        actual_concepts), f"Missing concepts in cash flow statement: {std_concepts - actual_concepts}"

    # Test statement of changes in equity
    equity = aapl_financials.get_statement_of_changes_in_equity(standard=True)
    assert equity is not None
    std_concepts = {concept.concept for concept in StatementOfChangesInEquity.concepts}
    actual_concepts = set(equity.data.reset_index()['concept'])
    assert std_concepts.issuperset(
        actual_concepts), f"Missing concepts in statement of changes in equity: {std_concepts - actual_concepts}"

    # Test statement of comprehensive income
    std_comprehensive_income = aapl_financials.get_statement_of_comprehensive_income(standard=True)
    assert std_comprehensive_income is not None
    std_concepts = {concept.concept for concept in StatementOfComprehensiveIncome.concepts}
    actual_concepts = set(std_comprehensive_income.data.reset_index()['concept'])
    assert std_concepts.issuperset(
        actual_concepts), f"Missing concepts in statement of comprehensive income: {std_concepts - actual_concepts}"

    # Test cover page
    std_cover_page = aapl_financials.get_cover_page()
    assert std_cover_page is not None
    # Cover page doesn't have standard concepts defined, so we just check if it exists


def test_statement_definition_durations(teradyne_xbrl):
    sd = teradyne_xbrl.get_statement_definition('CondensedConsolidatedStatementsOfCashFlows')
    assert sd.durations == {'instant', '6 months', '3 months', '1 month'}
    instance = teradyne_xbrl.instance
    durations = instance.facts.duration.value_counts().to_frame()
    print(durations)


def test_duration_with_data_selected_for_quarterly_income_statement(teradyne_xbrl):
    fin = Financials(teradyne_xbrl)
    cs = fin.get_cash_flow_statement()
    df = cs.get_dataframe()
    cols = df.columns.tolist()
    assert 'Jun 30, 2024' in cols
    assert 'Jul 02, 2023' in cols


def test_formatting_of_equity_value():
    filing = Filing(company='Palo Alto Networks Inc', cik=1327567, form='10-K', filing_date='2024-09-06', accession_no='0001327567-24-000029')
    xb = filing.xbrl()
    st = xb.get_statement('CONSOLIDATEDBALANCESHEETSParenthetical')
    assert st.name == 'CONSOLIDATEDBALANCESHEETSParenthetical'
